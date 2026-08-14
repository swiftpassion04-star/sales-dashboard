"""DB-free tests for the Dashboard sales-table XLSX export.

No pytest in this environment -- discovered via the repo's stdlib test_*
runner, same as every other tests/test_*.py file here.
"""

import ast
import sys
from datetime import date
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from ui.sales_export_ui import (
    SALES_EXPORT_HEADERS,
    build_sales_export_filename,
    build_sales_export_rows,
    build_sales_export_xlsx,
    sales_export_product_name,
)


DASHBOARD_SOURCE = (Path(__file__).resolve().parents[1] / "pages" / "dashboard.py").read_text(encoding="utf-8")

SAMPLE_ROWS = [
    {
        "sale_time": "09:52",
        "sale_type": "UPSELL",
        "order_id": "688194",
        "sku": "SP573",
        "product_name": "เลโก้นีม 50 ชิ้น - refill",
        "quantity": 1,
        "amount": 290.0,
        "created_staff": "จินดามณี คงมี (ครีม)",
    },
    {
        "sale_time": "15:06",
        "sale_type": "NEW_ORDER",
        "order_id": "690523",
        "sku": "SP323",
        "product_name": "แม่เหล็กรูปตัวL",
        "quantity": 2,
        "amount": 99.0,
        "created_staff": "จินดามณี คงมี (ครีม)",
    },
]


def test_headers_match_the_table_header_row():
    # The table's own header list lives in pages/dashboard.py; if someone
    # adds a column there this test fails rather than the export silently
    # drifting out of sync with what is on screen.
    header_line = next(
        line for line in DASHBOARD_SOURCE.splitlines() if line.strip().startswith("header = [")
    )
    table_headers = ast.literal_eval(header_line.split("header = ", 1)[1])
    assert table_headers == SALES_EXPORT_HEADERS, (table_headers, SALES_EXPORT_HEADERS)


def test_product_name_joins_sku_and_name():
    assert sales_export_product_name(SAMPLE_ROWS[0]) == "SP573 เลโก้นีม 50 ชิ้น - refill"


def test_product_name_handles_missing_parts():
    assert sales_export_product_name({"sku": "SP1", "product_name": ""}) == "SP1"
    assert sales_export_product_name({"sku": "", "product_name": "ชื่อ"}) == "ชื่อ"
    assert sales_export_product_name({}) == ""


def test_rows_are_numbered_from_one_in_display_order():
    exported = build_sales_export_rows(SAMPLE_ROWS)
    assert [r["ลำดับ"] for r in exported] == [1, 2]
    assert exported[0]["เลขคำสั่งซื้อ"] == "688194"
    assert exported[1]["ประเภทคำสั่งซื้อ"] == "NEW_ORDER"


def test_numeric_columns_are_real_numbers_not_display_strings():
    # ราคาอัพ must be a float so Excel can sum the column; the table shows
    # it as "290.00" but that formatting must not leak into the sheet.
    exported = build_sales_export_rows(SAMPLE_ROWS)
    assert isinstance(exported[0]["ราคาอัพ"], float) and exported[0]["ราคาอัพ"] == 290.0
    assert isinstance(exported[0]["จำนวนชิ้น"], int) and exported[1]["จำนวนชิ้น"] == 2
    assert isinstance(exported[0]["ลำดับ"], int)


def test_missing_and_malformed_values_do_not_raise():
    exported = build_sales_export_rows([{}, {"quantity": "x", "amount": None}])
    assert exported[0]["ชื่อสินค้า"] == ""
    assert exported[0]["จำนวนชิ้น"] == 0
    assert exported[1]["จำนวนชิ้น"] == 0
    assert exported[1]["ราคาอัพ"] == 0.0


def test_empty_rows_produce_no_export_rows():
    assert build_sales_export_rows([]) == []
    assert build_sales_export_rows(None) == []


def test_xlsx_is_a_real_workbook_with_the_expected_grid():
    data = build_sales_export_xlsx(SAMPLE_ROWS)
    assert data[:2] == b"PK", "xlsx must be a zip container"
    ws = openpyxl.load_workbook(BytesIO(data)).active
    assert ws.title == "sales_report"
    assert [c.value for c in ws[1]] == SALES_EXPORT_HEADERS
    assert ws.max_row == 3  # header + 2 rows
    assert ws.max_column == len(SALES_EXPORT_HEADERS)
    # Thai text survives the round-trip, and the amount stays numeric.
    assert ws.cell(row=2, column=5).value == "SP573 เลโก้นีม 50 ชิ้น - refill"
    assert ws.cell(row=2, column=7).value == 290.0
    assert isinstance(ws.cell(row=3, column=6).value, int)


def test_filename_includes_range_when_given():
    name = build_sales_export_filename(date(2026, 8, 1), date(2026, 8, 14))
    assert name.startswith("crm_sales_report_20260801_20260814_")
    assert name.endswith(".xlsx")
    assert build_sales_export_filename(None, None).startswith("crm_sales_report_")


def test_page_exports_display_rows_not_the_full_result_set():
    # The button must receive the limited `display_rows`, otherwise the file
    # would contain more than the table shows.
    assert "render_sales_report_download(display_rows" in DASHBOARD_SOURCE
    assert "render_sales_report_download(rows" not in DASHBOARD_SOURCE


def test_export_does_not_introduce_a_database_call():
    export_source = (Path(__file__).resolve().parents[1] / "ui" / "sales_export_ui.py").read_text(encoding="utf-8")
    for needle in ("neon_connection", "cur.execute", "select ", "fetch_sales_report"):
        assert needle not in export_source, needle
